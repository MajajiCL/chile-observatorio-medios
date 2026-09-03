# -*- coding: utf-8 -*-
"""
Incorpora el Censo de Poblacion y Vivienda 2024 (INE) al catalogo comunal.

Por que hace falta: Chile Abierto cubre economia, educacion, salud, seguridad y
gestion municipal, pero no dice nada de discapacidad, pueblos originarios,
hacinamiento, servicios basicos, migracion ni envejecimiento. Son justamente las
dimensiones donde suele estar el problema que ningun indicador economico muestra.

Nivel: comuna. La base manzana-entidad del Censo 2024 existe (189 variables) pero
el INE solo la publica desde su portal de geodatos, sin URL directa ni API, asi
que no se puede automatizar aqui. Si alguien la descarga a mano, el sitio ya esta
preparado para bajar de nivel.

Los archivos son planillas pensadas para leerse a ojo, no para maquinas: filas de
titulo, hojas multiples y una fila 'Pais' mezclada con las comunas. El parser
localiza la fila de encabezado por el texto 'Codigo comuna' y se queda solo con
las hojas donde cada comuna aparece una vez, que son las del agregado comunal.
"""
import json, os, re, sys, unicodedata, urllib.request

import openpyxl

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ORIG = os.path.join(BASE, "data", "censo_2024")
OUT = os.path.join(BASE, "data", "raw")
os.makedirs(ORIG, exist_ok=True)

INDICE = "https://censo2024.ine.gob.cl/estadisticas/"


def sin_tildes(s):
    return "".join(c for c in unicodedata.normalize("NFD", str(s))
                   if unicodedata.category(c) != "Mn").lower().strip()


def descargar():
    req = urllib.request.Request(INDICE, headers={"User-Agent": "observatorio-civico/1.0"})
    with urllib.request.urlopen(req, timeout=60) as r:
        html = r.read().decode("utf-8", errors="ignore")

    urls = sorted(set(re.findall(r'href="(https?://[^"]+\.xlsx)"', html, re.I)))
    print(f"{len(urls)} planillas publicadas en el sitio del Censo 2024")

    locales = []
    for u in urls:
        nombre = u.split("/")[-1]
        destino = os.path.join(ORIG, nombre)
        if not os.path.exists(destino):
            try:
                req = urllib.request.Request(u, headers={"User-Agent": "observatorio-civico/1.0"})
                with urllib.request.urlopen(req, timeout=180) as r, open(destino, "wb") as f:
                    f.write(r.read())
                print(f"   bajado  {nombre[:70]}")
            except Exception as e:
                print(f"   FALLO   {nombre[:60]}  {e}")
                continue
        locales.append(destino)
    return locales


def leer_hoja(ws):
    """Devuelve (encabezados, filas) si la hoja esta desagregada por comuna."""
    filas = list(ws.iter_rows(max_row=min(ws.max_row, 8000), values_only=True))
    hdr_i = None
    for i, row in enumerate(filas[:12]):
        textos = [sin_tildes(c) for c in row if c is not None]
        if any(t.startswith("codigo comuna") for t in textos):
            hdr_i = i
            break
    if hdr_i is None:
        return None, None

    hdr = [str(c).strip() if c is not None else "" for c in filas[hdr_i]]
    try:
        col_cc = next(i for i, h in enumerate(hdr) if sin_tildes(h).startswith("codigo comuna"))
    except StopIteration:
        return None, None

    datos, vistos = [], set()
    for row in filas[hdr_i + 1:]:
        if col_cc >= len(row):
            continue
        cc = row[col_cc]
        if not isinstance(cc, (int, float)):
            continue
        cc = str(int(cc))
        # En las regiones 1 a 9 el codigo trae 4 digitos (Iquique = 1101) y en
        # las demas 5 (Arica = 15101). Se normaliza a 5 con relleno, que es la
        # forma que usan las fuentes estadisticas y el resto del proyecto.
        if not (4 <= len(cc) <= 5):         # descarta tambien la fila 'Pais' (codigo 0)
            continue
        cc = cc.zfill(5)
        if cc in vistos:
            # la comuna se repite: la hoja tiene una dimension extra
            # (tramos de edad, tipologias). No sirve para un valor comunal.
            return None, None
        vistos.add(cc)
        datos.append((cc, row))
    return (hdr, datos) if len(datos) > 200 else (None, None)


def main():
    archivos = descargar()
    print()

    catalogo, series = [], {}

    for path in sorted(archivos):
        nombre = os.path.basename(path).replace(".xlsx", "")
        tema = re.sub(r"^[A-Z]\d+[-_]", "", nombre).replace("-", " ").replace("_", " ").strip()
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            print(f"  no se pudo abrir {nombre[:50]}: {e}")
            continue

        usadas = 0
        for hoja in wb.sheetnames:
            if sin_tildes(hoja).startswith("nota"):
                continue
            hdr, datos = leer_hoja(wb[hoja])
            if not datos:
                continue

            # Una hoja con decenas de columnas numericas no es una tabla de
            # indicadores sino una matriz: "Migracion interna" cruza cada comuna
            # de origen contra cada comuna de destino y genera 352 columnas que
            # no significan nada como indicador comunal.
            numericas = sum(
                1 for j in range(len(hdr))
                if sum(1 for _, row in datos[:40]
                       if j < len(row) and isinstance(row[j], (int, float))) > 30
            )
            if numericas > 40:
                continue

            omitir = ("codigo", "region", "provincia", "comuna", "")
            for j, h in enumerate(hdr):
                hs = sin_tildes(h)
                if not hs or any(hs.startswith(o) for o in omitir if o):
                    continue
                vals = {}
                for cc, row in datos:
                    v = row[j] if j < len(row) else None
                    if isinstance(v, (int, float)):
                        vals[cc] = float(v)
                if len(vals) < 200:
                    continue

                code = "censo24_" + re.sub(r"[^a-z0-9]+", "_",
                                           sin_tildes(nombre) + "_" + hs)[:58].strip("_")
                if code in series:
                    continue
                series[code] = vals
                catalogo.append({
                    "code": code,
                    "name_es": f"{h.strip()} — {tema}",
                    "unit": "%" if "%" in h or "porcentaje" in hs else "personas",
                    "category": "censo",
                    "direction": "ambiguous",
                    "year": 2024,
                    "national_avg": round(sum(vals.values()) / len(vals), 2),
                    "source": "INE — Censo de Población y Vivienda 2024",
                    "source_url": "https://censo2024.ine.gob.cl/estadisticas/",
                })
                usadas += 1
        if usadas:
            print(f"  {usadas:3} series  <-  {nombre[:64]}")

    print(f"\n{len(series)} series comunales extraidas del Censo 2024")

    for code, vals in series.items():
        meta = next(c for c in catalogo if c["code"] == code)
        with open(os.path.join(OUT, f"ind_{code}.json"), "w", encoding="utf-8") as f:
            json.dump({
                "indicator": meta,
                "values": [{"comuna_code": k, "value": v} for k, v in vals.items()],
                "total": len(vals),
            }, f, ensure_ascii=False, separators=(",", ":"))

    # fusionar en el catalogo que lee el motor
    p_ind = os.path.join(OUT, "indicators.json")
    d = json.load(open(p_ind, encoding="utf-8"))
    existentes = {r["code"] for r in d["data"]}
    nuevos = [c for c in catalogo if c["code"] not in existentes]
    d["data"].extend(nuevos)
    with open(p_ind, "w", encoding="utf-8") as f:
        json.dump(d, f, ensure_ascii=False, separators=(",", ":"))
    print(f"catalogo: {len(d['data'])} indicadores en total (+{len(nuevos)})")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
